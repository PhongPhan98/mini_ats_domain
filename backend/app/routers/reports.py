import csv
import hashlib
import hmac
import io
import json
import smtplib
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from urllib import request

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import settings
from app.database import get_db
from app.models import Candidate
from app.rbac import require_roles
from app.routers.analytics import summary as analytics_summary

router = APIRouter(prefix="/api/reports", tags=["reports"])

SCHEDULES_FILE = Path(__file__).resolve().parents[1] / "data" / "report_schedules.json"


def _ensure_schedules_file():
    SCHEDULES_FILE.parent.mkdir(parents=True, exist_ok=True)
    if not SCHEDULES_FILE.exists():
        SCHEDULES_FILE.write_text(json.dumps({"schedules": []}, indent=2), encoding="utf-8")


def _load_schedules() -> dict:
    _ensure_schedules_file()
    return json.loads(SCHEDULES_FILE.read_text(encoding="utf-8"))


def _save_schedules(payload: dict):
    _ensure_schedules_file()
    SCHEDULES_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _normalize_status(value: str | None) -> str:
    if not value:
        return "applied"
    v = value.strip().lower()
    return {"new": "applied", "shortlisted": "screening"}.get(v, v)


def _candidate_rows(db: Session):
    rows = list(db.execute(select(Candidate)).scalars().all())
    return [
        {
            "id": c.id,
            "name": c.name or "",
            "email": c.email or "",
            "phone": c.phone or "",
            "status": _normalize_status(c.status),
            "years_of_experience": c.years_of_experience or "",
            "skills": ", ".join(c.skills or []),
            "source": (c.parsed_json or {}).get("source", "direct"),
            "created_at": c.created_at.isoformat() if c.created_at else "",
        }
        for c in rows
    ]


def _build_xlsx_bytes(db: Session) -> bytes:
    analytics = analytics_summary(db=db, _=None)
    candidates = _candidate_rows(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"
    ws.append(["id", "name", "email", "phone", "status", "years_of_experience", "skills", "source", "created_at"])
    for r in candidates:
        ws.append([
            r["id"],
            r["name"],
            r["email"],
            r["phone"],
            r["status"],
            r["years_of_experience"],
            r["skills"],
            r["source"],
            r["created_at"],
        ])

    ws2 = wb.create_sheet("Summary")
    ws2.append(["metric", "value"])
    ws2.append(["total_candidates", analytics.total_candidates])
    ws2.append(["hired_count", analytics.hired_count])
    ws2.append(["avg_time_to_hire_days", analytics.avg_time_to_hire_days])

    ws3 = wb.create_sheet("Conversions")
    ws3.append(["stage", "rate_pct"])
    for x in analytics.conversion_rates:
        ws3.append([x["stage"], x["rate_pct"]])

    ws4 = wb.create_sheet("Sources")
    ws4.append(["source", "count", "share_pct"])
    for x in analytics.source_effectiveness:
        ws4.append([x["source"], x["count"], x["share_pct"]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_pdf_bytes(db: Session) -> bytes:
    analytics = analytics_summary(db=db, _=None)

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    _, height = A4

    y = height - 40
    c.setFont("Helvetica-Bold", 16)
    c.drawString(40, y, "Mini ATS Recruitment Report")
    y -= 24
    c.setFont("Helvetica", 10)
    c.drawString(40, y, f"Generated at: {datetime.utcnow().isoformat()} UTC")

    y -= 28
    c.setFont("Helvetica-Bold", 12)
    c.drawString(40, y, "Highlights")
    y -= 18
    c.setFont("Helvetica", 10)
    c.drawString(50, y, f"Total candidates: {analytics.total_candidates}")
    y -= 14
    c.drawString(50, y, f"Hired count: {analytics.hired_count}")
    y -= 14
    c.drawString(50, y, f"Average time-to-hire (days): {analytics.avg_time_to_hire_days}")

    y -= 24
    c.setFont("Helvetica-Bold", 12)
    c.drawString(40, y, "Conversion Rates")
    y -= 18
    c.setFont("Helvetica", 10)
    for row in analytics.conversion_rates:
        c.drawString(50, y, f"{row['stage']}: {row['rate_pct']}%")
        y -= 14
        if y < 60:
            c.showPage()
            y = height - 40

    c.save()
    return buf.getvalue()


def _send_email_with_attachments(to_email: str, subject: str, body: str, attachments: list[tuple[str, bytes, str]]):
    if not settings.smtp_enabled:
        return False, "smtp_disabled"
    if not settings.smtp_host or not settings.smtp_from_email:
        return False, "smtp_incomplete_config"

    try:
        msg = MIMEMultipart()
        msg["Subject"] = subject
        msg["From"] = settings.smtp_from_email
        msg["To"] = to_email
        msg.attach(MIMEText(body, "plain", "utf-8"))

        for filename, data, mime_subtype in attachments:
            part = MIMEApplication(data, _subtype=mime_subtype)
            part.add_header("Content-Disposition", "attachment", filename=filename)
            msg.attach(part)

        with smtplib.SMTP(settings.smtp_host, settings.smtp_port, timeout=12) as server:
            if settings.smtp_use_tls:
                server.starttls()
            if settings.smtp_username:
                server.login(settings.smtp_username, settings.smtp_password)
            server.sendmail(settings.smtp_from_email, [to_email], msg.as_string())

        return True, "email_sent"
    except Exception as e:
        return False, f"email_error:{e}"


def _post_webhook(url: str, payload: dict):
    raw = json.dumps(payload).encode("utf-8")
    headers = {"Content-Type": "application/json"}
    if settings.webhook_signing_secret:
        sig = hmac.new(settings.webhook_signing_secret.encode("utf-8"), raw, hashlib.sha256).hexdigest()
        headers["X-MiniATS-Signature"] = f"sha256={sig}"

    req = request.Request(url, data=raw, headers=headers, method="POST")
    with request.urlopen(req, timeout=10) as resp:
        return resp.status


@router.get("/candidates.csv")
def export_candidates_csv(
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin", "recruiter", "hiring_manager")),
):
    rows = _candidate_rows(db)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["id", "name", "email", "phone", "status", "years_of_experience", "skills", "source", "created_at"])
    for r in rows:
        writer.writerow([r["id"], r["name"], r["email"], r["phone"], r["status"], r["years_of_experience"], r["skills"], r["source"], r["created_at"]])

    filename = f"mini_ats_candidates_{datetime.utcnow().date().isoformat()}.csv"
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/analytics.csv")
def export_analytics_csv(
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin", "recruiter", "hiring_manager")),
):
    analytics = analytics_summary(db=db, _=None)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["section", "key", "value"])
    writer.writerow(["overview", "total_candidates", analytics.total_candidates])
    writer.writerow(["overview", "hired_count", analytics.hired_count])
    writer.writerow(["overview", "avg_time_to_hire_days", analytics.avg_time_to_hire_days])
    for x in analytics.conversion_rates:
        writer.writerow(["conversion", x["stage"], x["rate_pct"]])

    filename = f"mini_ats_analytics_{datetime.utcnow().date().isoformat()}.csv"
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/reports.xlsx")
def export_reports_xlsx(
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin", "recruiter", "hiring_manager")),
):
    data = _build_xlsx_bytes(db)
    filename = f"mini_ats_reports_{datetime.utcnow().date().isoformat()}.xlsx"
    return StreamingResponse(io.BytesIO(data), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/report.pdf")
def export_report_pdf(
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin", "recruiter", "hiring_manager")),
):
    data = _build_pdf_bytes(db)
    filename = f"mini_ats_report_{datetime.utcnow().date().isoformat()}.pdf"
    return StreamingResponse(io.BytesIO(data), media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/pdf-snapshot")
def export_pdf_snapshot_json(
    db: Session = Depends(get_db),
    _=Depends(require_roles("admin", "recruiter", "hiring_manager")),
):
    analytics = analytics_summary(db=db, _=None)
    return {
        "title": "Mini ATS Recruitment Report",
        "generated_at": datetime.utcnow().isoformat(),
        "highlights": {
            "total_candidates": analytics.total_candidates,
            "hired_count": analytics.hired_count,
            "avg_time_to_hire_days": analytics.avg_time_to_hire_days,
        },
        "charts": {
            "funnel": analytics.status_distribution,
            "conversion_rates": analytics.conversion_rates,
            "source_effectiveness": analytics.source_effectiveness,
            "top_skills": analytics.top_skills,
        },
    }


@router.get("/schedules")
def list_report_schedules(_=Depends(require_roles("admin", "hiring_manager"))):
    return _load_schedules()


@router.post("/schedules")
def upsert_report_schedule(payload: dict, _=Depends(require_roles("admin", "hiring_manager"))):
    data = _load_schedules()
    schedules = data.get("schedules", [])

    schedule_id = payload.get("id") or f"schedule-{int(datetime.utcnow().timestamp())}"
    row = {
        "id": schedule_id,
        "enabled": bool(payload.get("enabled", True)),
        "cadence": payload.get("cadence", "weekly"),
        "weekday": payload.get("weekday", "monday"),
        "time_utc": payload.get("time_utc", "02:00"),
        "delivery": payload.get("delivery", {"mode": "email", "to": ""}),
        "formats": payload.get("formats", ["pdf", "xlsx"]),
        "updated_at": datetime.utcnow().isoformat(),
    }

    replaced = False
    for i, s in enumerate(schedules):
        if s.get("id") == schedule_id:
            schedules[i] = row
            replaced = True
            break
    if not replaced:
        schedules.append(row)

    data["schedules"] = schedules
    _save_schedules(data)
    return row


@router.post("/schedules/{schedule_id}/run")
def run_report_schedule(schedule_id: str, db: Session = Depends(get_db), _=Depends(require_roles("admin", "hiring_manager"))):
    data = _load_schedules()
    schedule = next((s for s in data.get("schedules", []) if s.get("id") == schedule_id), None)
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")

    formats = schedule.get("formats", ["pdf", "xlsx"])
    attachments = []
    if "pdf" in formats:
        attachments.append((f"report_{datetime.utcnow().date().isoformat()}.pdf", _build_pdf_bytes(db), "pdf"))
    if "xlsx" in formats:
        attachments.append((f"report_{datetime.utcnow().date().isoformat()}.xlsx", _build_xlsx_bytes(db), "vnd.openxmlformats-officedocument.spreadsheetml.sheet"))

    delivery = schedule.get("delivery", {})
    mode = delivery.get("mode", "email")

    if mode == "email":
        to = delivery.get("to")
        if not to:
            raise HTTPException(status_code=400, detail="Missing delivery.to email")
        ok, result = _send_email_with_attachments(
            to,
            subject="Mini ATS Weekly Report",
            body="Attached are your latest ATS reports.",
            attachments=attachments,
        )
        return {"ok": ok, "result": result, "schedule_id": schedule_id}

    if mode == "webhook":
        url = delivery.get("to")
        if not url:
            raise HTTPException(status_code=400, detail="Missing delivery.to webhook URL")
        status = _post_webhook(
            url,
            {
                "schedule_id": schedule_id,
                "generated_at": datetime.utcnow().isoformat(),
                "formats": formats,
                "note": "Binary attachments are email-only in v1; webhook receives metadata snapshot.",
                "summary": analytics_summary(db=db, _=None).model_dump(),
            },
        )
        return {"ok": True, "result": f"webhook:{status}", "schedule_id": schedule_id}

    raise HTTPException(status_code=400, detail=f"Unsupported delivery mode '{mode}'")
